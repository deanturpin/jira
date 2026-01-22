"""Generate Excel reports with velocity and epic timeline data."""

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Any


class ExcelGenerator:
    """Generates formatted Excel reports with charts."""

    def __init__(self, filename: str):
        self.filename = filename
        self.workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

    def _apply_header_style(self, sheet, row: int, num_cols: int):
        """Apply consistent header styling."""
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            bottom=Side(style='thin', color='000000')
        )

        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _auto_size_columns(self, sheet):
        """Auto-size columns based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def add_velocity_sheet(self, velocity_data: list[dict[str, Any]], stats: dict[str, float]):
        """Add velocity history sheet with chart."""
        sheet = self.workbook.create_sheet('Velocity History')

        # Headers
        headers = ['Sprint', 'Start Date', 'End Date', 'Committed Points', 'Completed Points', 'Completion Rate']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        self._apply_header_style(sheet, 1, len(headers))

        # Data rows
        for row, sprint in enumerate(velocity_data, start=2):
            sheet.cell(row=row, column=1, value=sprint['sprint_name'])
            sheet.cell(row=row, column=2, value=sprint.get('start_date', ''))
            sheet.cell(row=row, column=3, value=sprint.get('end_date', ''))
            sheet.cell(row=row, column=4, value=sprint['total_points'])
            sheet.cell(row=row, column=5, value=sprint['completed_points'])
            sheet.cell(row=row, column=6, value=f"{sprint['completion_rate']:.1%}")

        # Add statistics section
        stats_start_row = len(velocity_data) + 4
        sheet.cell(row=stats_start_row, column=1, value='Velocity Statistics')
        sheet.cell(row=stats_start_row, column=1).font = Font(bold=True, size=12)

        stats_data = [
            ('Average Velocity', stats['mean']),
            ('Median Velocity', stats['median']),
            ('Standard Deviation', stats['std_dev']),
            ('Minimum', stats['min']),
            ('Maximum', stats['max'])
        ]

        for i, (label, value) in enumerate(stats_data, start=1):
            sheet.cell(row=stats_start_row + i, column=1, value=label)
            sheet.cell(row=stats_start_row + i, column=2, value=f"{value:.1f}")

        # Create velocity trend chart
        chart = LineChart()
        chart.title = 'Sprint Velocity Trend'
        chart.y_axis.title = 'Story Points'
        chart.x_axis.title = 'Sprint'

        data = Reference(sheet, min_col=5, min_row=1, max_row=len(velocity_data) + 1)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=len(velocity_data) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        sheet.add_chart(chart, f'H2')

        self._auto_size_columns(sheet)

    def add_epic_timeline_sheet(
        self,
        timeline: list[dict[str, Any]],
        sprint_capacity: list[dict[str, Any]]
    ):
        """Add epic timeline sheet with Gantt-style visualisation."""
        sheet = self.workbook.create_sheet('Epic Timeline')

        # Headers
        headers = [
            'Epic Key', 'Epic Name', 'Total Points', 'Completed',
            'Remaining', 'Start Sprint', 'End Sprint', 'Duration (Sprints)', 'Status'
        ]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        self._apply_header_style(sheet, 1, len(headers))

        # Data rows
        for row, epic in enumerate(timeline, start=2):
            sheet.cell(row=row, column=1, value=epic['epic_key'])
            sheet.cell(row=row, column=2, value=epic['epic_name'])
            sheet.cell(row=row, column=3, value=epic['total_points'])
            sheet.cell(row=row, column=4, value=epic['completed_points'])
            sheet.cell(row=row, column=5, value=epic['remaining_points'])
            sheet.cell(row=row, column=6, value=epic.get('start_sprint', 'N/A'))
            sheet.cell(row=row, column=7, value=epic.get('end_sprint', 'N/A'))
            sheet.cell(row=row, column=8, value=epic.get('sprint_count', 0))
            sheet.cell(row=row, column=9, value=epic['status'])

            # Colour code by status
            status_cell = sheet.cell(row=row, column=9)
            if epic['status'] == 'scheduled':
                status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Create chart showing remaining points per epic
        chart = BarChart()
        chart.type = 'bar'
        chart.title = 'Epic Remaining Work'
        chart.y_axis.title = 'Epic'
        chart.x_axis.title = 'Story Points'

        data = Reference(sheet, min_col=5, min_row=1, max_row=len(timeline) + 1)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=len(timeline) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        sheet.add_chart(chart, f'K2')

        self._auto_size_columns(sheet)

    def add_capacity_planning_sheet(
        self,
        sprint_projections: list[dict[str, Any]],
        sprint_capacity: list[dict[str, Any]]
    ):
        """Add capacity planning sheet showing future sprint allocations."""
        sheet = self.workbook.create_sheet('Capacity Planning')

        # Headers
        headers = [
            'Sprint #', 'Start Date', 'End Date',
            'Projected Capacity', 'Allocated Points', 'Available Capacity'
        ]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        self._apply_header_style(sheet, 1, len(headers))

        # Data rows
        for row, sprint in enumerate(sprint_capacity, start=2):
            allocated = sprint['capacity'] - sprint['remaining_capacity']
            sheet.cell(row=row, column=1, value=f"Sprint {sprint['sprint_number']}")
            sheet.cell(row=row, column=2, value=sprint['start_date'][:10])
            sheet.cell(row=row, column=3, value=sprint['end_date'][:10])
            sheet.cell(row=row, column=4, value=sprint['capacity'])
            sheet.cell(row=row, column=5, value=allocated)
            sheet.cell(row=row, column=6, value=sprint['remaining_capacity'])

            # Colour code based on capacity utilisation
            utilisation = allocated / sprint['capacity'] if sprint['capacity'] > 0 else 0
            capacity_cell = sheet.cell(row=row, column=6)

            if utilisation >= 0.9:
                capacity_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif utilisation >= 0.7:
                capacity_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                capacity_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        # Add details of epic allocations
        details_start_row = len(sprint_capacity) + 4
        sheet.cell(row=details_start_row, column=1, value='Epic Allocations by Sprint')
        sheet.cell(row=details_start_row, column=1).font = Font(bold=True, size=12)

        detail_row = details_start_row + 2
        for sprint in sprint_capacity:
            if sprint['assigned_epics']:
                sheet.cell(row=detail_row, column=1, value=f"Sprint {sprint['sprint_number']}")
                sheet.cell(row=detail_row, column=1).font = Font(bold=True)
                detail_row += 1

                for epic in sprint['assigned_epics']:
                    sheet.cell(row=detail_row, column=2, value=epic['epic_key'])
                    sheet.cell(row=detail_row, column=3, value=epic['epic_name'])
                    sheet.cell(row=detail_row, column=4, value=epic['points_allocated'])
                    detail_row += 1

                detail_row += 1

        self._auto_size_columns(sheet)

    def add_summary_sheet(self, velocity_data: list[dict[str, Any]], stats: dict[str, float]):
        """Add summary overview sheet."""
        sheet = self.workbook.create_sheet('Summary', 0)

        # Title
        sheet.cell(row=1, column=1, value='Jira Planning Report')
        sheet.cell(row=1, column=1).font = Font(bold=True, size=16)

        # Generation timestamp
        sheet.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sheet.cell(row=2, column=1).font = Font(italic=True)

        # Key metrics
        metrics_start = 4
        sheet.cell(row=metrics_start, column=1, value='Key Metrics')
        sheet.cell(row=metrics_start, column=1).font = Font(bold=True, size=14)

        metrics = [
            ('Average Velocity (Story Points/Sprint)', f"{stats['mean']:.1f}"),
            ('Velocity Standard Deviation', f"{stats['std_dev']:.1f}"),
            ('Number of Historical Sprints Analysed', len(velocity_data)),
            ('Most Recent Sprint Velocity', f"{velocity_data[-1]['completed_points']:.1f}" if velocity_data else 'N/A')
        ]

        for i, (label, value) in enumerate(metrics, start=1):
            sheet.cell(row=metrics_start + i, column=1, value=label)
            sheet.cell(row=metrics_start + i, column=2, value=value)
            sheet.cell(row=metrics_start + i, column=1).font = Font(bold=True)

        # Instructions
        instructions_start = metrics_start + len(metrics) + 3
        sheet.cell(row=instructions_start, column=1, value='Sheet Descriptions')
        sheet.cell(row=instructions_start, column=1).font = Font(bold=True, size=14)

        descriptions = [
            ('Velocity History', 'Historical sprint velocity data with trend chart'),
            ('Epic Timeline', 'Projected completion dates for epics based on capacity'),
            ('Capacity Planning', 'Future sprint capacity allocation and availability')
        ]

        for i, (sheet_name, description) in enumerate(descriptions, start=1):
            sheet.cell(row=instructions_start + i, column=1, value=sheet_name)
            sheet.cell(row=instructions_start + i, column=2, value=description)
            sheet.cell(row=instructions_start + i, column=1).font = Font(bold=True)

        self._auto_size_columns(sheet)

    def save(self):
        """Save the workbook to file."""
        self.workbook.save(self.filename)
        return self.filename
